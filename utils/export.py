"""
Export Module
Handles conversion of Pin data to Pandas DataFrames and Excel files.

Future extensions:
    - Supabase / PostgreSQL push
    - Google Sheets export
    - CSV / JSON export
    - Scheduled export (APScheduler integration)
"""

import io
import logging
import re
from datetime import datetime
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from scraper.pinterest_scraper import Pin

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions — single source of truth for ordering & renaming
# ---------------------------------------------------------------------------

COLUMNS = [
    "keyword",
    "position",
    "score",
    "title",
    "description",
    "pin_url",
    "image_url",
]

COLUMN_DISPLAY_NAMES = {
    "keyword": "Keyword",
    "position": "Position",
    "score": "Score",
    "title": "Title",
    "description": "Description",
    "pin_url": "Pin URL",
    "image_url": "Image URL",
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def pins_to_dataframe(pins: list["Pin"]) -> pd.DataFrame:
    """
    Convert a list of Pin objects into a cleaned, analysis-ready DataFrame.

    Steps:
        1. Flatten dataclass fields to dict rows
        2. Select & order columns
        3. Clean missing values
        4. Remove duplicate pin URLs
        5. Sort by score descending
    """
    if not pins:
        return pd.DataFrame(columns=COLUMNS)

    rows = [
        {
            "keyword": p.keyword,
            "position": p.position,
            "score": p.score,
            "title": p.title or "",
            "description": p.description or "",
            "pin_url": p.pin_url or "",
            "image_url": p.image_url or "",
        }
        for p in pins
    ]

    df = pd.DataFrame(rows, columns=COLUMNS)

    # --- Clean ---
    df["title"] = df["title"].fillna("").str.strip()
    df["description"] = df["description"].fillna("").str.strip()
    df["pin_url"] = df["pin_url"].fillna("").str.strip()
    df["image_url"] = df["image_url"].fillna("").str.strip()

    # --- Deduplicate on pin URL (keep first / highest ranked occurrence) ---
    df = df.drop_duplicates(subset=["pin_url"], keep="first")

    # --- Sort ---
    df = df.sort_values("score", ascending=False).reset_index(drop=True)

    return df


def dataframe_to_excel_bytes(df: pd.DataFrame, keyword: str) -> bytes:
    """
    Convert a DataFrame to a styled Excel workbook and return raw bytes.
    The bytes can be streamed directly by Streamlit's download_button.
    """
    buffer = io.BytesIO()

    # Write base Excel via pandas (fast)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_display = df.rename(columns=COLUMN_DISPLAY_NAMES)
        df_display.to_excel(writer, index=False, sheet_name="Pinterest Pins")

        _style_worksheet(writer.sheets["Pinterest Pins"], df_display)

        # --- Metadata sheet ---
        meta_df = pd.DataFrame(
            {
                "Property": ["Keyword", "Total Pins", "Exported At", "Tool"],
                "Value": [
                    keyword,
                    len(df),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Pinterest Niche Scraper",
                ],
            }
        )
        meta_df.to_excel(writer, index=False, sheet_name="Metadata")
        _style_metadata_sheet(writer.sheets["Metadata"])

    return buffer.getvalue()


def safe_filename(keyword: str) -> str:
    """
    Generate a safe file name from a keyword.
    Example: "fitness motivation" → "pinterest_scraping_fitness_motivation.xlsx"
    """
    slug = re.sub(r"[^\w\s-]", "", keyword).strip()
    slug = re.sub(r"[\s-]+", "_", slug)
    return f"pinterest_scraping_{slug}.xlsx"


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="E60023")   # Pinterest red
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill("solid", fgColor="FFF5F5")  # Very light red tint
_THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

COLUMN_WIDTHS = {
    "Keyword": 18,
    "Position": 10,
    "Score": 10,
    "Title": 40,
    "Description": 55,
    "Pin URL": 50,
    "Image URL": 50,
}


def _style_worksheet(ws, df: pd.DataFrame) -> None:
    """Apply header styles, column widths, and alternating row colours."""
    # Header row
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = _ALT_ROW_FILL if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER

    # Column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        width = COLUMN_WIDTHS.get(col_name, 20)
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _style_metadata_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 30


# ---------------------------------------------------------------------------
# Future: Supabase integration stub
# ---------------------------------------------------------------------------

def push_to_supabase(df: pd.DataFrame, table: str = "pinterest_pins") -> None:
    """
    Placeholder for Supabase integration.
    Future implementation:
        from supabase import create_client
        client = create_client(SUPABASE_URL, SUPABASE_KEY)
        client.table(table).upsert(df.to_dict("records")).execute()
    """
    raise NotImplementedError("Supabase integration not yet configured.")